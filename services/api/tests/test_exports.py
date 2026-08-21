import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.auth import get_current_user
from app.main import app
from app.repositories.brands import InMemoryBrandsRepository, get_brands_repository
from app.repositories.calculations import InMemoryCalculationsRepository, get_calculations_repository
from app.repositories.exports import InMemoryExportsRepository, get_exports_repository
from app.repositories.mapping_templates import InMemoryMappingTemplatesRepository, get_mapping_templates_repository
from app.repositories.matches import InMemoryMatchesRepository, get_matches_repository
from app.repositories.projects import InMemoryProjectsRepository, get_projects_repository
from app.repositories.ratings import InMemoryRatingsRepository, get_ratings_repository
from app.repositories.uploads import InMemoryUploadsRepository, get_uploads_repository

from ._auth_helpers import make_test_user


def _returning(instance):
    def _get():
        return instance

    return _get


@pytest.fixture
def client():
    overrides = {
        get_projects_repository: InMemoryProjectsRepository(),
        get_brands_repository: InMemoryBrandsRepository(),
        get_uploads_repository: InMemoryUploadsRepository(),
        get_ratings_repository: InMemoryRatingsRepository(),
        get_matches_repository: InMemoryMatchesRepository(),
        get_calculations_repository: InMemoryCalculationsRepository(),
        get_mapping_templates_repository: InMemoryMappingTemplatesRepository(),
        get_exports_repository: InMemoryExportsRepository(),
    }
    for dependency, instance in overrides.items():
        app.dependency_overrides[dependency] = _returning(instance)
    app.dependency_overrides[get_current_user] = _returning(make_test_user(role='owner'))
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


def _project(client):
    return client.post('/api/projects', json={'projectName': 'Export Test', 'client': 'Client X'}).json()


def _seed_and_calculate(client, project_id):
    brands = client.get(f'/api/projects/{project_id}/brands').json()
    brand_id = brands[0]['brandId'] if brands else client.post(
        f'/api/projects/{project_id}/brands', json={'name': 'Brand A'}
    ).json()['brandId']

    # AIT/Morning Show matches; XYZ FM/Drive Time has nothing in ratings -> unmatched.
    csv = (
        b'Channel,Programme,Day,Spots\n'
        b'AIT,Morning Show,Monday,3\n'
        b'XYZ FM,Drive Time,Wednesday,1\n'
    )
    client.post(
        f'/api/projects/{project_id}/uploads',
        data={'kind': 'brand_report', 'brand_id': brand_id, 'default_medium': 'TV'},
        files={'file': ('report.csv', io.BytesIO(csv), 'text/csv')},
    )
    dataset = client.post(
        '/api/ratings-datasets',
        json={'provider': 'P', 'rows': [{'medium': 'TV', 'station': 'AIT', 'day': 'MON', 'programme': 'Morning Show', 'rating': 5.0}]},
    ).json()
    client.post(f"/api/projects/{project_id}/ratings-datasets/{dataset['ratingsDatasetId']}/attach")
    return client.post(f'/api/projects/{project_id}/calculate').json()


def test_create_export_defaults_to_latest_run(client):
    project = _project(client)
    run = _seed_and_calculate(client, project['projectId'])

    response = client.post(f"/api/projects/{project['projectId']}/exports", json={})
    assert response.status_code == 201
    body = response.json()
    assert body['runId'] == run['runId']
    assert body['format'] == 'xlsx_full'
    assert body['downloadUrl'] == f"/api/projects/{project['projectId']}/exports/{body['exportId']}/download"


def test_create_export_with_no_run_yet(client):
    project = _project(client)
    response = client.post(f"/api/projects/{project['projectId']}/exports", json={})
    assert response.status_code == 201
    assert response.json()['runId'] is None


def test_create_export_rejects_unsupported_format(client):
    project = _project(client)
    response = client.post(f"/api/projects/{project['projectId']}/exports", json={'format': 'pdf'})
    assert response.status_code == 422


def test_create_export_rejects_unknown_run_id(client):
    project = _project(client)
    response = client.post(f"/api/projects/{project['projectId']}/exports", json={'runId': 'does-not-exist'})
    assert response.status_code == 404


def test_export_freezes_the_run_it_was_created_for(client):
    project = _project(client)
    project_id = project['projectId']
    first_run = _seed_and_calculate(client, project_id)

    export = client.post(f'/api/projects/{project_id}/exports', json={}).json()
    assert export['runId'] == first_run['runId']

    second_run = _seed_and_calculate(client, project_id)
    assert second_run['runId'] != first_run['runId']

    # Re-fetching the export still shows the original run, not the new latest.
    refetched = client.get(f"/api/projects/{project_id}/exports/{export['exportId']}").json()
    assert refetched['runId'] == first_run['runId']


def test_list_and_get_exports(client):
    project = _project(client)
    project_id = project['projectId']
    _seed_and_calculate(client, project_id)
    created = client.post(f'/api/projects/{project_id}/exports', json={}).json()

    listed = client.get(f'/api/projects/{project_id}/exports').json()
    assert len(listed) == 1
    assert listed[0]['exportId'] == created['exportId']

    fetched = client.get(f"/api/projects/{project_id}/exports/{created['exportId']}")
    assert fetched.status_code == 200
    assert fetched.json() == created


def test_get_missing_export_returns_404(client):
    project = _project(client)
    response = client.get(f"/api/projects/{project['projectId']}/exports/does-not-exist")
    assert response.status_code == 404


def test_download_returns_a_readable_workbook_with_expected_sheets(client):
    project = _project(client)
    project_id = project['projectId']
    _seed_and_calculate(client, project_id)
    export = client.post(f'/api/projects/{project_id}/exports', json={}).json()

    response = client.get(f"/api/projects/{project_id}/exports/{export['exportId']}/download")
    assert response.status_code == 200
    assert response.headers['content-type'].startswith('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    assert 'Export_Test_export.xlsx' in response.headers['content-disposition']

    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == [
        'Project Info', 'GRP & SOV', 'Brand Comparison', 'Station Performance', 'Programme Performance',
        'Spot-Level Calculations', 'Unmatched Records', 'Ratings Used',
    ]

    project_info = workbook['Project Info']
    values = {row[0].value: row[1].value for row in project_info.iter_rows(min_row=2)}
    assert values['Project'] == 'Export Test'
    assert values['Client'] == 'Client X'
    assert values['Total Brands'] == 1
    assert values['Total Spend'] == 0

    grp_sov = workbook['GRP & SOV']
    header = [cell.value for cell in grp_sov[1]]
    assert header == [
        'Brand', 'Total GRPs', 'TV GRPs', 'Cable TV GRPs', 'Radio GRPs', 'SOV %', 'Spots', 'Avg Rating', 'Spend', 'SOE %',
    ]
    first_data_row = [cell.value for cell in grp_sov[2]]
    assert first_data_row[0] == 'Brand A'
    assert first_data_row[8] == 0  # Spend — no cost column mapped in this fixture's upload

    unmatched = workbook['Unmatched Records']
    unmatched_rows = list(unmatched.iter_rows(min_row=2, values_only=True))
    assert len(unmatched_rows) == 1
    assert unmatched_rows[0][2] == 'XYZ FM'  # Station column


def test_download_with_no_run_yet_still_produces_a_workbook(client):
    project = _project(client)
    export = client.post(f"/api/projects/{project['projectId']}/exports", json={}).json()
    response = client.get(f"/api/projects/{project['projectId']}/exports/{export['exportId']}/download")
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    values = {row[0].value: row[1].value for row in workbook['Project Info'].iter_rows(min_row=2)}
    assert values['Run'] == 'No calculation has been run for this project yet.'


def test_download_missing_export_returns_404(client):
    project = _project(client)
    response = client.get(f"/api/projects/{project['projectId']}/exports/does-not-exist/download")
    assert response.status_code == 404


def test_exports_scoped_to_project(client):
    project_a = _project(client)
    project_b = client.post('/api/projects', json={'projectName': 'Other'}).json()
    _seed_and_calculate(client, project_a['projectId'])
    created = client.post(f"/api/projects/{project_a['projectId']}/exports", json={}).json()

    response = client.get(f"/api/projects/{project_b['projectId']}/exports/{created['exportId']}")
    assert response.status_code == 404
